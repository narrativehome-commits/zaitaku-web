from flask import Flask, render_template
from datetime import datetime
import openpyxl
import os

app = Flask(__name__)

EXCEL_FILE = "data/在宅医療充実体制加算カウント用.xlsx"
SHEET_NAME = "R8.4"

def get_summary():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME]
    col = 5

    def safe_percent(val):
        if val is None:
            return "---"
        if isinstance(val, float) and val <= 1.0:
            return f"{val:.0%}"
        return str(val)

    def safe_int(val):
        if val is None:
            return "---"
        try:
            return str(int(val)) + " 人"
        except:
            return str(val)

    def safe_str(val):
        if val is None:
            return "---"
        return str(val)

    return {
        "全体人数":        safe_int(ws.cell(row=233, column=col).value),
        "該当者数":        safe_int(ws.cell(row=234, column=col).value),
        "割合":            safe_percent(ws.cell(row=235, column=col).value),
        "認知症Ⅳ・M人数":  safe_int(ws.cell(row=236, column=col).value),
        "認知症割合":      safe_percent(ws.cell(row=237, column=col).value),
        "基準":            safe_str(ws.cell(row=239, column=col).value),
        "判定":            safe_str(ws.cell(row=240, column=col).value),
    }

@app.route("/")
def index():
    try:
        summary = get_summary()
        error = None
    except Exception as e:
        summary = None
        error = str(e)
    now = datetime.now().strftime("%Y年%m月%d日 %H:%M 現在")
    return render_template("index.html", summary=summary, now=now,
                           sheet=SHEET_NAME, error=error)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
